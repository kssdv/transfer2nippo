from openpyxl import load_workbook

def read_excel(path):
    gyomu = []
    syokan = []

    # Excelファイルを読み込む
    workbook = load_workbook(path)

    # シート名をリストで取得
    sheet_names = workbook.sheetnames

    sheet = workbook[sheet_names[0]]
    # sheet = sheet_names[0]
    # print(sheet_names[0])  # 例: ['Sheet1', 'Sheet2', 'Sheet3']

    # cell_value = sheet[0]

    # A列の1〜5行目の値を取得 「業務内容」
    gyomu.append("【業務内容】")
    for row in range(5, 29):
        cell = sheet[f"A{row}"]
        if cell.value is not None:
            gyomu.append(cell.value)
            print(f"A{row}: {cell.value}")
    print("---------------")
    # 所感
    syokan.append("【本日の所感】")
    for row in range(30, 37):
        cell = sheet[f"A{row}"]
        if cell.value is not None:
            syokan.append(cell.value)
            print(f"A{row}: {cell.value}")
    
    return [gyomu, syokan]

if __name__ == "__main__":
    pass
